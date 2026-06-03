"""
Export the live "Performance Creative Repository" Google Sheet into committed,
diffable JSON snapshots under experiments/.

This is the versioned MIRROR of the operational registry (Option 2):
the Google Sheet stays the editing surface; this repo holds the audit trail.

Refresh workflow:
  1. Export the Sheet to xlsx (via the google-workspace MCP / File > Download),
     e.g. the file lands in C:\\Users\\nikhi\\.workspace-mcp\\attachments\\...xlsx
  2. Run:  python scripts/export_registry_to_json.py "<path-to-export.xlsx>"
  3. Review `git diff experiments/` and commit.

Source spreadsheet: June 1 - Performance Creative Repository
ID: 1ay8EBJ3fRjpx0U80ErEqMwJGXlY0ynh_FzTfGcTGXaA

Only the four operational tabs are mirrored (Brief / Creative / Experiment /
Status). The Context & Legend tab is reference doctrine, not operational data,
and is intentionally excluded.
"""
import json
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

SOURCE_SPREADSHEET_ID = "1ay8EBJ3fRjpx0U80ErEqMwJGXlY0ynh_FzTfGcTGXaA"
SOURCE_TITLE = "June 1 - Performance Creative Repository"

# Sheet-tab name -> output JSON filename (only operational tabs are mirrored).
TAB_TO_FILE = {
    "0_Brief_Registry": "brief_registry.json",
    "1_Creative_Registry": "creative_registry.json",
    "2_Experiment_Registry": "experiment_registry.json",
    "3_Status_Tracker": "status_tracker.json",
}

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = REPO_ROOT / "experiments"


def norm(value):
    """Coerce a cell value to a stable, diff-friendly string."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def rows_to_records(ws):
    """First non-empty row = header. Each later non-empty row -> dict keyed by header."""
    grid = list(ws.iter_rows(values_only=True))
    if not grid:
        return []
    header = [norm(c) for c in grid[0]]
    # trim trailing empty header columns
    while header and header[-1] == "":
        header.pop()
    if not header:
        return []
    records = []
    for raw in grid[1:]:
        cells = [norm(c) for c in raw][: len(header)]
        cells += [""] * (len(header) - len(cells))
        if all(c == "" for c in cells):
            continue  # skip fully-empty rows
        records.append({header[i]: cells[i] for i in range(len(header))})
    return records


def main():
    if len(sys.argv) < 2:
        sys.exit(
            "Usage: python scripts/export_registry_to_json.py <path-to-sheet-export.xlsx>"
        )
    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        sys.exit(f"ERROR: file not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    OUT_DIR.mkdir(exist_ok=True)

    tab_counts = {}
    for tab_name, out_file in TAB_TO_FILE.items():
        if tab_name not in wb.sheetnames:
            print(f"WARN: tab '{tab_name}' not found in workbook; skipping.")
            continue
        records = rows_to_records(wb[tab_name])
        (OUT_DIR / out_file).write_text(
            json.dumps(records, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
        tab_counts[tab_name] = len(records)
        print(f"  {tab_name:<22} -> experiments/{out_file}  ({len(records)} records)")

    meta = {
        "_note": "Generated mirror of the live Google Sheet. DO NOT hand-edit; "
                 "edit the Sheet and re-run scripts/export_registry_to_json.py.",
        "source_spreadsheet_id": SOURCE_SPREADSHEET_ID,
        "source_title": SOURCE_TITLE,
        "exported_from_xlsx": xlsx_path.name,
        "export_timestamp": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "tab_record_counts": tab_counts,
    }
    (OUT_DIR / "_snapshot_meta.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(f"  snapshot meta          -> experiments/_snapshot_meta.json")
    print("Done.")


if __name__ == "__main__":
    main()
